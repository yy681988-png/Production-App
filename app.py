import streamlit as st
import pandas as pd
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import plotly.express as px
import datetime
import io
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from google import genai
    GEMINI_AVAILABLE = True
except ImportError:
    GEMINI_AVAILABLE = False

SHEET_KEY = "17y_KBs5xQqTY_63UtMC22Sxru7X9jxhg86LvM1WL9us"

# ---------------------------------------------------------------------------
# Connexion (mise en cache avec cache_resource pour éviter de se reconnecter
# à Google à chaque interaction / rerun de Streamlit)
# ---------------------------------------------------------------------------
@st.cache_resource
def get_client():
    creds_dict = st.secrets["gcp_service_account"]
    scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
    return gspread.authorize(creds)


@st.cache_resource
def get_sheet():
    try:
        return get_client().open_by_key(SHEET_KEY)
    except Exception as e:
        st.error(f"Impossible de se connecter à Google Sheets : {e}")
        st.stop()


sheet = get_sheet()


def retry_on_quota(func, *args, max_retries=4, **kwargs):
    """Exécute un appel à l'API Google Sheets avec retry + backoff exponentiel
    en cas d'erreur 429 (quota dépassé)."""
    delay = 2
    last_error = None
    for attempt in range(max_retries):
        try:
            return func(*args, **kwargs)
        except gspread.exceptions.APIError as e:
            status = None
            try:
                status = e.response.status_code
            except Exception:
                pass
            if status == 429:
                last_error = e
                time.sleep(delay)
                delay *= 2
                continue
            raise
    raise last_error


@st.cache_data(ttl=60)
def get_df(sheet_name):
    try:
        worksheet = retry_on_quota(sheet.worksheet, sheet_name)
        data = retry_on_quota(worksheet.get_all_values)
    except gspread.exceptions.APIError as e:
        st.error(
            f"Quota Google Sheets dépassé pour la feuille '{sheet_name}'. "
            "Réessayez dans quelques instants, ou augmentez le quota "
            "'Read requests per minute' dans Google Cloud Console."
        )
        return pd.DataFrame()
    except Exception as e:
        st.error(f"Erreur de lecture de la feuille '{sheet_name}' : {e}")
        return pd.DataFrame()

    if not data:
        return pd.DataFrame()

    df = pd.DataFrame(data[1:], columns=data[0])
    if 'ref' in df.columns:
        df['ref'] = df['ref'].astype(str).str.replace("'", "", regex=False)
    return df


def ensure_columns(df, columns):
    """Garantit que les colonnes attendues existent, même si la feuille est vide."""
    if df.empty:
        return pd.DataFrame(columns=columns)
    for col in columns:
        if col not in df.columns:
            df[col] = None
    return df


def save_to_sheet(sheet_name, df):
    try:
        worksheet = retry_on_quota(sheet.worksheet, sheet_name)
        retry_on_quota(worksheet.clear)
        retry_on_quota(worksheet.update, [df.columns.values.tolist()] + df.values.tolist())
        # On ne vide que le cache de CETTE feuille, pas de toutes les feuilles :
        # évite de forcer une relecture inutile de tout le classeur à chaque écriture.
        get_df.clear(sheet_name)
    except gspread.exceptions.APIError as e:
        st.error(
            f"Quota Google Sheets dépassé lors de l'enregistrement dans '{sheet_name}'. "
            "Réessayez dans quelques instants."
        )
    except Exception as e:
        st.error(f"Erreur lors de l'enregistrement dans '{sheet_name}' : {e}")


def safe_drop(df, idx, sheet_name):
    """Supprime une ligne par index seulement si l'index existe réellement."""
    if idx in df.index:
        save_to_sheet(sheet_name, df.drop(idx).reset_index(drop=True))
        st.rerun()
    else:
        st.error("Index invalide : aucune ligne correspondante trouvée.")


def get_monthly_trends(selected_refs=None):
    """Calcule l'évolution mensuelle Objectif vs Production, en s'appuyant sur
    la date des logs de production et le mois du plan mensuel."""
    df_logs = ensure_columns(get_df("production_logs"), ['date', 'ref', 'qty'])
    df_plan = ensure_columns(get_df("monthly_plan"), ['month', 'ref', 'target', 'price'])

    df_logs['qty'] = pd.to_numeric(df_logs['qty'], errors='coerce').fillna(0)
    df_plan['target'] = pd.to_numeric(df_plan['target'], errors='coerce').fillna(0)

    if not df_logs.empty:
        df_logs['date_parsed'] = pd.to_datetime(df_logs['date'], errors='coerce')
        df_logs['month'] = df_logs['date_parsed'].dt.strftime('%Y-%m')

    if selected_refs:
        if not df_logs.empty:
            df_logs = df_logs[df_logs['ref'].isin(selected_refs)]
        if not df_plan.empty:
            df_plan = df_plan[df_plan['ref'].isin(selected_refs)]

    if not df_logs.empty and 'month' in df_logs.columns:
        monthly_actual = df_logs.groupby('month')['qty'].sum().reset_index().rename(columns={'qty': 'Production'})
    else:
        monthly_actual = pd.DataFrame(columns=['month', 'Production'])

    if not df_plan.empty:
        monthly_target = df_plan.groupby('month')['target'].sum().reset_index().rename(columns={'target': 'Objectif'})
    else:
        monthly_target = pd.DataFrame(columns=['month', 'Objectif'])

    trends = pd.merge(monthly_target, monthly_actual, on='month', how='outer').fillna(0)
    if not trends.empty:
        trends = trends.dropna(subset=['month'])
        trends = trends.sort_values('month')
        for col in ['Objectif', 'Production']:
            trends[col] = pd.to_numeric(trends[col], errors='coerce').fillna(0)
        trends['Taux (%)'] = (trends['Production'] / trends['Objectif'].replace(0, 1) * 100)
        trends.loc[trends['Objectif'] == 0, 'Taux (%)'] = 0

    return trends


def get_dashboard_data():
    """Calcule les données comparatives (plan vs réalisé) une seule fois,
    réutilisées par l'onglet Dashboard et l'onglet Rapports & IA."""
    df_products = get_df("products")
    df_logs = ensure_columns(get_df("production_logs"), ['date', 'ref', 'qty'])
    df_plan = ensure_columns(get_df("monthly_plan"), ['month', 'ref', 'target', 'price'])

    df_logs['qty'] = pd.to_numeric(df_logs['qty'], errors='coerce').fillna(0)
    df_plan['target'] = pd.to_numeric(df_plan['target'], errors='coerce').fillna(0)
    df_plan['price'] = pd.to_numeric(df_plan['price'], errors='coerce').fillna(0)

    prod_grouped = df_logs.groupby('ref')['qty'].sum().reset_index() if not df_logs.empty else pd.DataFrame(columns=['ref', 'qty'])
    plan_grouped = df_plan.groupby('ref').agg(target=('target', 'sum'), price=('price', 'mean')).reset_index() if not df_plan.empty else pd.DataFrame(columns=['ref', 'target', 'price'])

    df_compare = pd.merge(plan_grouped, prod_grouped, on='ref', how='outer').fillna(0)
    for col in ['target', 'qty', 'price']:
        if col in df_compare.columns:
            df_compare[col] = pd.to_numeric(df_compare[col], errors='coerce').fillna(0)

    if not df_compare.empty:
        df_compare['Taux (%)'] = (df_compare['qty'] / df_compare['target'].replace(0, 1) * 100)
        df_compare.loc[df_compare['target'] == 0, 'Taux (%)'] = 0
        df_compare['CA réalisé'] = df_compare['qty'] * df_compare['price']

    return df_products, df_plan, df_logs, df_compare


# ---------------------------------------------------------------------------
# Export Excel professionnel
# ---------------------------------------------------------------------------
def write_styled_sheet(wb, title, df, numeric_cols=None):
    ws = wb.create_sheet(title)
    numeric_cols = numeric_cols or []

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    if df is None or df.empty:
        ws.append(["Aucune donnée disponible"])
        ws["A1"].font = Font(name="Arial", italic=True, color="888888")
        return ws

    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for _, row in df.iterrows():
        ws.append(list(row))

    for r_idx in range(2, ws.max_row + 1):
        for c_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = Font(name="Arial", size=10)
            cell.border = border
            if col_name in numeric_cols:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")

    for c_idx, col_name in enumerate(df.columns, start=1):
        values = [str(col_name)] + [str(v) for v in df[col_name].tolist()]
        max_len = max(len(v) for v in values)
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"
    return ws


def generate_excel_report(df_products, df_plan, df_logs, df_compare):
    wb = Workbook()
    wb.remove(wb.active)

    write_styled_sheet(wb, "Produits", df_products)
    write_styled_sheet(wb, "Plan Mensuel", df_plan, numeric_cols=["target", "price"])
    write_styled_sheet(wb, "Production", df_logs, numeric_cols=["qty"])
    write_styled_sheet(wb, "Dashboard", df_compare, numeric_cols=["target", "qty", "price", "Taux (%)", "CA réalisé"])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ---------------------------------------------------------------------------
# Analyse par IA (Claude, via l'API Anthropic)
# ---------------------------------------------------------------------------
def build_data_summary(df_compare):
    total_target = df_compare['target'].sum()
    total_qty = df_compare['qty'].sum()
    perf = (total_qty / total_target * 100) if total_target > 0 else 0

    lines = [
        f"Objectif total : {total_target:.0f}",
        f"Production totale réalisée : {total_qty:.0f}",
        f"Performance globale : {perf:.1f}%",
        "",
        "Détail par référence (ref, objectif, produit, taux %, CA réalisé) :",
    ]
    for _, row in df_compare.iterrows():
        lines.append(
            f"- {row['ref']} : objectif={row['target']:.0f}, produit={row['qty']:.0f}, "
            f"taux={row['Taux (%)']:.1f}%, CA={row.get('CA réalisé', 0):.2f}"
        )
    return "\n".join(lines)


def get_ai_analysis(data_summary):
    if not GEMINI_AVAILABLE:
        return None, "Le package 'google-genai' n'est pas installé. Ajoutez-le à requirements.txt."

    api_key = st.secrets.get("GEMINI_API_KEY")
    if not api_key:
        return None, "Clé GEMINI_API_KEY absente des secrets Streamlit (st.secrets)."

    try:
        client = genai.Client(api_key=api_key)
        prompt = (
            "Tu es un analyste de production industrielle. Voici les données de "
            "performance du mois (objectifs vs production réelle) :\n\n"
            f"{data_summary}\n\n"
            "Rédige une analyse concise en français, structurée avec des puces :\n"
            "1) Synthèse générale de la performance\n"
            "2) Références en retard à surveiller en priorité\n"
            "3) Références en avance / bonnes performances\n"
            "4) 2 à 3 recommandations concrètes et actionnables"
        )
        response = client.models.generate_content(
            model="gemini-3.5-flash",
            contents=prompt,
        )
        return response.text, None
    except Exception as e:
        return None, f"Erreur lors de l'appel à l'IA : {e}"


st.set_page_config(layout="wide")
st.title("Gestion de Production Pro")

if st.sidebar.button("Actualiser les données (Refresh)"):
    st.cache_data.clear()
    st.rerun()

tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs(["Produits", "Plan", "Saisie", "Dashboard", "Rapports & IA", "Tendances"])

# ---------------------------------------------------------------------------
# 1. Produits
# ---------------------------------------------------------------------------
with tab1:
    st.header("Gestion des Produits")
    c1, c2 = st.columns(2)
    ref_p = c1.text_input("Référence", key="p_ref")
    name_p = c2.text_input("Nom", key="p_name")

    if st.button("Ajouter Produit"):
        if not ref_p.strip():
            st.error("La référence ne peut pas être vide.")
        else:
            df = ensure_columns(get_df("products"), ['ref', 'name'])
            new_row = pd.DataFrame({'ref': [ref_p], 'name': [name_p]})
            df = pd.concat([df, new_row], ignore_index=True).drop_duplicates(subset='ref', keep='last')
            save_to_sheet("products", df)
            st.rerun()

    df_products = get_df("products")
    st.dataframe(df_products)

    product_refs = df_products['ref'].tolist() if not df_products.empty and 'ref' in df_products.columns else []
    del_p = st.selectbox("Choisir Ref à supprimer", product_refs, key="del_prod")
    if st.button("Supprimer Produit"):
        if del_p:
            save_to_sheet("products", df_products[df_products['ref'] != del_p])
            st.rerun()
        else:
            st.warning("Aucun produit à supprimer.")

# ---------------------------------------------------------------------------
# 2. Plan mensuel
# ---------------------------------------------------------------------------
with tab2:
    st.header("Plan Mensuel")
    df_prods = get_df("products")
    product_list = df_prods['ref'].tolist() if not df_prods.empty and 'ref' in df_prods.columns else []

    m, r, t, p = st.columns(4)
    sel_date = m.date_input("Mois", value=datetime.date.today())
    month_val = sel_date.strftime("%Y-%m")
    ref = r.selectbox("Ref", options=product_list, key="plan_r")
    target = t.number_input("Target", value=0, key="plan_t")
    price = p.number_input("Price", value=0, key="plan_p")

    if st.button("Ajouter Plan"):
        if not ref:
            st.error("Veuillez choisir une référence.")
        else:
            df = ensure_columns(get_df("monthly_plan"), ['month', 'ref', 'target', 'price'])
            # Empêche les doublons pour le même mois + la même référence :
            # si une ligne existe déjà, on la met à jour au lieu de l'additionner.
            mask = (df['month'] == month_val) & (df['ref'] == ref)
            if mask.any():
                df.loc[mask, ['target', 'price']] = [float(target), float(price)]
                st.info("Une entrée existait déjà pour ce mois/référence : elle a été mise à jour.")
            else:
                new_row = pd.DataFrame({'month': [month_val], 'ref': [ref], 'target': [float(target)], 'price': [float(price)]})
                df = pd.concat([df, new_row], ignore_index=True)
            save_to_sheet("monthly_plan", df)
            st.rerun()

    df_plan = get_df("monthly_plan")
    st.dataframe(df_plan)

    idx = st.number_input("Indice ligne à supprimer (Plan)", min_value=0, step=1, key="plan_del")
    if st.button("Supprimer ligne Plan"):
        safe_drop(df_plan, idx, "monthly_plan")

# ---------------------------------------------------------------------------
# 3. Saisie de production
# ---------------------------------------------------------------------------
with tab3:
    st.header("Saisie de Production")
    df_prods = get_df("products")
    product_list = df_prods['ref'].tolist() if not df_prods.empty and 'ref' in df_prods.columns else []

    d, r, q = st.columns(3)
    date = d.date_input("Date", key="log_d")
    ref = r.selectbox("Ref", options=product_list, key="log_r")
    qty = q.number_input("Qty", value=0, key="log_q")

    if st.button("Ajouter Log"):
        if not ref:
            st.error("Veuillez choisir une référence.")
        else:
            df = ensure_columns(get_df("production_logs"), ['date', 'ref', 'qty'])
            new_row = pd.DataFrame({'date': [str(date)], 'ref': [ref], 'qty': [float(qty)]})
            df = pd.concat([df, new_row], ignore_index=True)
            save_to_sheet("production_logs", df)
            st.rerun()

    df_logs = get_df("production_logs")
    st.dataframe(df_logs)

    idx_l = st.number_input("Indice ligne à supprimer (Log)", min_value=0, step=1, key="log_del")
    if st.button("Supprimer ligne Log"):
        safe_drop(df_logs, idx_l, "production_logs")

# ---------------------------------------------------------------------------
# 4. Dashboard
# ---------------------------------------------------------------------------
with tab4:
    st.header("Tableau de Bord Comparatif")
    _, _, _, df_compare = get_dashboard_data()

    if df_compare.empty:
        st.info("Aucune donnée disponible pour le moment.")
    else:
        ref_search = st.multiselect("Filtrer par Ref", df_compare['ref'].unique(), key="dash_search")
        if ref_search:
            df_compare = df_compare[df_compare['ref'].isin(ref_search)]

        c1, c2, c3 = st.columns(3)
        c1.metric("Objectif Total", int(df_compare['target'].sum()))
        c2.metric("Production Totale", int(df_compare['qty'].sum()))

        total_target = df_compare['target'].sum()
        total_qty = df_compare['qty'].sum()
        total_perc = (total_qty / total_target * 100) if total_target > 0 else 0
        c3.metric("Performance Globale", f"{total_perc:.1f}%")

        st.divider()
        st.subheader("Détail par Produit")
        st.dataframe(
            df_compare.style.background_gradient(subset=['Taux (%)'], cmap='RdYlGn'),
            column_config={
                "target": st.column_config.NumberColumn("Target", format="%.2f"),
                "qty": st.column_config.NumberColumn("Qty", format="%.2f"),
                "Taux (%)": st.column_config.NumberColumn("Taux (%)", format="%.2f"),
                "CA réalisé": st.column_config.NumberColumn("CA réalisé", format="%.2f"),
            },
            use_container_width=True
        )

        fig = px.bar(
            df_compare,
            x='ref',
            y=['target', 'qty'],
            barmode='group',
            title="Performance par Référence",
            text_auto='.2s',
            color_discrete_map={"target": "#3498db", "qty": "#2ecc71"}
        )
        fig.update_xaxes(type='category')
        fig.update_layout(template="plotly_dark", legend_title="Indicateur")
        st.plotly_chart(fig, use_container_width=True)

# ---------------------------------------------------------------------------
# 5. Rapports & IA
# ---------------------------------------------------------------------------
with tab5:
    st.header("Rapports & Analyse IA")
    df_products_r, df_plan_r, df_logs_r, df_compare_r = get_dashboard_data()

    st.subheader("Résumé rapide")
    if df_compare_r.empty:
        st.info("Pas encore de données à résumer.")
    else:
        total_target = df_compare_r['target'].sum()
        total_qty = df_compare_r['qty'].sum()
        perf = (total_qty / total_target * 100) if total_target > 0 else 0

        c1, c2, c3 = st.columns(3)
        c1.metric("Objectif Total", f"{total_target:.0f}")
        c2.metric("Production Totale", f"{total_qty:.0f}")
        c3.metric("Performance Globale", f"{perf:.1f}%")

        colw, colb = st.columns(2)
        with colw:
            st.markdown("**⚠️ Références en retard**")
            worst = df_compare_r.sort_values("Taux (%)").head(3)
            st.dataframe(worst[['ref', 'target', 'qty', 'Taux (%)']], use_container_width=True, hide_index=True)
        with colb:
            st.markdown("**✅ Meilleures performances**")
            best = df_compare_r.sort_values("Taux (%)", ascending=False).head(3)
            st.dataframe(best[['ref', 'target', 'qty', 'Taux (%)']], use_container_width=True, hide_index=True)

    st.divider()
    st.subheader("Export Excel")
    st.caption("Génère un classeur avec 4 feuilles (Produits, Plan Mensuel, Production, Dashboard), mis en forme automatiquement.")

    if st.button("📊 Générer le rapport Excel"):
        with st.spinner("Génération du fichier Excel..."):
            excel_buffer = generate_excel_report(df_products_r, df_plan_r, df_logs_r, df_compare_r)
            st.session_state["excel_report"] = excel_buffer.getvalue()

    if "excel_report" in st.session_state:
        st.download_button(
            label="📥 Télécharger le rapport Excel",
            data=st.session_state["excel_report"],
            file_name=f"rapport_production_{datetime.date.today()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    st.divider()
    st.subheader("Analyse par Intelligence Artificielle")
    st.caption(
        "Nécessite le package 'google-genai' dans requirements.txt et une clé "
        "GEMINI_API_KEY définie dans les secrets Streamlit (st.secrets)."
    )

    if st.button("🤖 Générer une analyse IA"):
        if df_compare_r.empty:
            st.warning("Pas assez de données pour générer une analyse.")
        else:
            with st.spinner("Analyse en cours..."):
                summary = build_data_summary(df_compare_r)
                analysis, error = get_ai_analysis(summary)
            if error:
                st.error(error)
            else:
                st.session_state["ai_analysis"] = analysis

    if "ai_analysis" in st.session_state:
        st.markdown(st.session_state["ai_analysis"])

# ---------------------------------------------------------------------------
# 6. Tendances Historiques
# ---------------------------------------------------------------------------
with tab6:
    st.header("Tendances Historiques")
    st.caption("Évolution de l'objectif et de la production réalisée, mois par mois.")

    df_products_t = get_df("products")
    product_list_t = df_products_t['ref'].tolist() if not df_products_t.empty and 'ref' in df_products_t.columns else []

    refs_filter = st.multiselect(
        "Filtrer par référence (laisser vide = toutes les références)",
        product_list_t,
        key="trend_refs"
    )
    trends = get_monthly_trends(refs_filter if refs_filter else None)

    if trends.empty:
        st.info("Pas encore assez de données pour afficher une tendance. Ajoutez des entrées dans 'Plan' et 'Saisie' sur plusieurs mois.")
    else:
        fig_trend = px.line(
            trends,
            x='month',
            y=['Objectif', 'Production'],
            markers=True,
            title="Évolution Objectif vs Production par mois",
            color_discrete_map={"Objectif": "#3498db", "Production": "#2ecc71"}
        )
        fig_trend.update_layout(
            template="plotly_dark",
            legend_title="Indicateur",
            xaxis_title="Mois",
            yaxis_title="Quantité"
        )
        fig_trend.update_xaxes(type='category')
        st.plotly_chart(fig_trend, use_container_width=True)

        fig_perf = px.bar(
            trends,
            x='month',
            y='Taux (%)',
            title="Taux de réalisation par mois (%)",
            text_auto='.1f',
            color_discrete_sequence=["#2ecc71"]
        )
        fig_perf.update_layout(template="plotly_dark", xaxis_title="Mois", yaxis_title="Taux (%)")
        fig_perf.update_xaxes(type='category')
        fig_perf.add_hline(y=100, line_dash="dash", line_color="red", annotation_text="Objectif 100%")
        st.plotly_chart(fig_perf, use_container_width=True)

        st.subheader("Détail mensuel")
        st.dataframe(
            trends,
            column_config={
                "month": st.column_config.TextColumn("Mois"),
                "Objectif": st.column_config.NumberColumn("Objectif", format="%.2f"),
                "Production": st.column_config.NumberColumn("Production", format="%.2f"),
                "Taux (%)": st.column_config.NumberColumn("Taux (%)", format="%.2f"),
            },
            use_container_width=True,
            hide_index=True
        )
